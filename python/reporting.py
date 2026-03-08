"""
reporting.py
Stage 4: Report Generation
Queries the SQLite database and produces a formatted Excel report
with multiple tabs: Summary, Fund Detail, Validation Issues, and Ingestion Log.
"""

import sqlite3
import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import DB_PATH, REPORT_TITLE, REPORT_PERIOD, REPORT_DATE, LP_NAME

logger = logging.getLogger(__name__)

DARK_BLUE = "1B2A4A"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
GREEN = "27AE60"
RED = "E74C3C"

header_font = Font(name="Arial", bold=True, color=WHITE, size=10)
header_fill = PatternFill("solid", fgColor=DARK_BLUE)
data_font = Font(name="Arial", size=10, color="404040")
title_font = Font(name="Arial", bold=True, color=DARK_BLUE, size=14)
subtitle_font = Font(name="Arial", bold=True, color=DARK_BLUE, size=11)
border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def generate_report(db_path=None, output_dir="output"):
    """Generate the full quarterly report as Excel."""
    if db_path is None:
        db_path = DB_PATH

    os.makedirs(output_dir, exist_ok=True)
    conn = sqlite3.connect(db_path)

    wb = Workbook()

    build_summary_sheet(wb, conn)
    build_fund_detail_sheet(wb, conn)
    build_validation_sheet(wb, conn)
    build_ingestion_log_sheet(wb, conn)

    conn.close()

    output_path = os.path.join(output_dir, f"Quarterly_Report_{REPORT_PERIOD.replace(' ', '_')}.xlsx")
    wb.save(output_path)
    logger.info(f"Report saved to {output_path}")

    return output_path


def build_summary_sheet(wb, conn):
    """Build the summary dashboard sheet."""
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = DARK_BLUE

    # Title
    ws.cell(row=1, column=1, value=REPORT_TITLE).font = title_font
    ws.cell(row=2, column=1, value=f"Period: {REPORT_PERIOD} | As of: {REPORT_DATE}").font = subtitle_font
    ws.cell(row=3, column=1, value=f"Prepared for: {LP_NAME}").font = data_font

    # Fetch summary stats
    stats = conn.execute("""
        SELECT
            COUNT(DISTINCT fund_name) as fund_count,
            COUNT(DISTINCT gp_name) as gp_count,
            COUNT(*) as investment_count,
            COALESCE(SUM(commitment_eur), 0) as total_commitment,
            COALESCE(SUM(called_eur), 0) as total_called,
            COALESCE(SUM(distributed_eur), 0) as total_distributed
        FROM investments
    """).fetchone()

    fund_count, gp_count, inv_count, total_commit, total_called, total_dist = stats

    # KPIs
    kpis = [
        ("Funds", fund_count),
        ("General Partners", gp_count),
        ("Investments", inv_count),
        ("Total Commitments (EUR)", total_commit),
        ("Total Called Capital (EUR)", total_called),
        ("Total Distributions (EUR)", total_dist),
        ("Call Rate", total_called / total_commit if total_commit > 0 else 0),
        ("DPI (Distributions / Called)", total_dist / total_called if total_called > 0 else 0),
    ]

    row = 5
    ws.cell(row=row, column=1, value="Portfolio Overview").font = subtitle_font
    row += 1

    for label, value in kpis:
        ws.cell(row=row, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        cell = ws.cell(row=row, column=2)
        cell.font = Font(name="Arial", size=10, bold=True, color=DARK_BLUE)

        if isinstance(value, float) and value < 10:
            cell.value = value
            cell.number_format = "0.00x" if "DPI" in label else "0.0%"
        elif isinstance(value, (int, float)) and value > 1000:
            cell.value = value
            cell.number_format = "#,##0"
        else:
            cell.value = value

        row += 1

    # By GP breakdown
    row += 1
    ws.cell(row=row, column=1, value="Breakdown by General Partner").font = subtitle_font
    row += 1

    gp_headers = ["GP Name", "Investments", "Commitment (EUR)", "Called (EUR)", "Distributed (EUR)", "Call Rate"]
    for c, h in enumerate(gp_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    row += 1

    gp_data = conn.execute("""
        SELECT gp_name, COUNT(*), SUM(commitment_eur), SUM(called_eur), SUM(distributed_eur)
        FROM investments GROUP BY gp_name ORDER BY SUM(commitment_eur) DESC
    """).fetchall()

    for gp_row in gp_data:
        gp_name, count, commit, called, dist = gp_row
        call_rate = called / commit if commit and commit > 0 else 0
        values = [gp_name, count, commit, called, dist, call_rate]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = data_font
            cell.border = border
            if c in [3, 4, 5]:
                cell.number_format = "#,##0"
            if c == 6:
                cell.number_format = "0.0%"
        row += 1

    # By vintage breakdown
    row += 1
    ws.cell(row=row, column=1, value="Breakdown by Vintage Year").font = subtitle_font
    row += 1

    v_headers = ["Vintage", "Investments", "Commitment (EUR)", "Called (EUR)", "Status Mix"]
    for c, h in enumerate(v_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    row += 1

    vintage_data = conn.execute("""
        SELECT vintage_year, COUNT(*), SUM(commitment_eur), SUM(called_eur),
            GROUP_CONCAT(DISTINCT status)
        FROM investments WHERE vintage_year IS NOT NULL
        GROUP BY vintage_year ORDER BY vintage_year
    """).fetchall()

    for v_row in vintage_data:
        for c, val in enumerate(v_row, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = data_font
            cell.border = border
            if c in [3, 4]:
                cell.number_format = "#,##0"
        row += 1

    # Auto-width
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 22


def build_fund_detail_sheet(wb, conn):
    """Build the fund-level detail sheet."""
    ws = wb.create_sheet("Fund Detail")
    ws.sheet_properties.tabColor = "2E5090"

    ws.cell(row=1, column=1, value="Investment Detail").font = title_font

    headers = [
        "Investment ID", "Company", "Fund", "GP", "Vintage", "Strategy", "Status",
        "Commitment (EUR)", "Called (EUR)", "Distributed (EUR)", "Call Rate", "DPI",
        "Original CCY", "Source File"
    ]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    data = conn.execute("""
        SELECT investment_id, company_name, fund_name, gp_name, vintage_year,
               strategy, status, commitment_eur, called_eur, distributed_eur,
               original_currency, source_file
        FROM investments ORDER BY gp_name, company_name
    """).fetchall()

    for r, row_data in enumerate(data, 4):
        inv_id, company, fund, gp, vintage, strategy, status, commit, called, dist, orig_ccy, source = row_data
        call_rate = (called or 0) / commit if commit and commit > 0 else 0
        dpi = (dist or 0) / called if called and called > 0 else 0

        values = [inv_id, company, fund, gp, vintage, strategy, status,
                  commit, called, dist, call_rate, dpi, orig_ccy, source]

        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = border
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
            if c in [8, 9, 10]:
                cell.number_format = "#,##0"
            if c == 11:
                cell.number_format = "0.0%"
            if c == 12:
                cell.number_format = "0.00x"

    ws.auto_filter.ref = f"A3:N{3 + len(data)}"
    ws.freeze_panes = "A4"

    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 18


def build_validation_sheet(wb, conn):
    """Build the validation issues sheet."""
    ws = wb.create_sheet("Validation Issues")
    ws.sheet_properties.tabColor = RED

    ws.cell(row=1, column=1, value="Data Quality Report").font = title_font

    headers = ["#", "Investment ID", "Check Type", "Field", "Issue", "Severity", "Source File"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    issues = conn.execute("""
        SELECT id, investment_id, check_type, field, issue, severity, source_file
        FROM validation_issues ORDER BY severity DESC, check_type
    """).fetchall()

    for r, row_data in enumerate(issues, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = border
            if c == 6:
                if val == "Critical":
                    cell.font = Font(name="Arial", size=10, color=RED, bold=True)
                elif val == "Warning":
                    cell.font = Font(name="Arial", size=10, color="F39C12", bold=True)

    # Summary at bottom
    r = 4 + len(issues) + 1
    total = len(issues)
    critical = sum(1 for i in issues if i[5] == "Critical")
    warning = sum(1 for i in issues if i[5] == "Warning")

    ws.cell(row=r, column=1, value="Summary").font = subtitle_font
    ws.cell(row=r + 1, column=1, value="Total Issues").font = data_font
    ws.cell(row=r + 1, column=2, value=total).font = data_font
    ws.cell(row=r + 2, column=1, value="Critical").font = data_font
    ws.cell(row=r + 2, column=2, value=critical).font = Font(name="Arial", size=10, color=RED)
    ws.cell(row=r + 3, column=1, value="Warning").font = data_font
    ws.cell(row=r + 3, column=2, value=warning).font = Font(name="Arial", size=10, color="F39C12")

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 22


def build_ingestion_log_sheet(wb, conn):
    """Build the ingestion log sheet."""
    ws = wb.create_sheet("Ingestion Log")
    ws.sheet_properties.tabColor = "404040"

    ws.cell(row=1, column=1, value="File Ingestion Log").font = title_font

    headers = ["#", "Filename", "GP Source", "Rows Ingested", "Columns", "Processed At"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    logs = conn.execute("SELECT * FROM ingestion_log ORDER BY id").fetchall()

    for r, row_data in enumerate(logs, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = border

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 25
